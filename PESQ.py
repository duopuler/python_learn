import numpy as np
from openpyxl import Workbook                  #导入excel插件
from openpyxl.utils import get_column_letter
import re    #正则表达式,用于检索关键字
import os    #文件操作
for num in [21,22,23,24,25,26]:    #文件编号list
    key_SNR_0dB = "0dB"            #关键字们
    key_SNR_5dB = "5dB"
    key_SNR_10dB = "10dB"
    key_HF = "HF"
    key_HS = "HS"

    list_HF_0dB = []              #属性们
    list_HF_5dB = []
    list_HF_10dB = []
    list_HF_alldB = []

    list_HS_0dB = []
    list_HS_5dB = []
    list_HS_10dB = []
    list_HS_alldB = []

    list_0dB = []
    list_5dB = []
    list_10dB = []
    list_alldB = []

    dirpath ="C:/Users/samsung/Desktop/python/" + str(num) +"/OUTPUT1/"     #根目录
    allfile = os.listdir(dirpath)    #列出根目录下所有文件
    pattern = re.compile(r'.*(.txt)') #生成正则表达式的pattern实例,表示任意txt文件名*.txt

    for onefile in allfile:          #遍历根目录
        match = pattern.match(onefile)   #匹配根目录符合文件
        # child = os.path.join("%s%s" %(dirpath,onefile))
        if match:                     #将符合匹配文件存入file_txt
            file_txt = onefile


    path_read = "C:/Users/samsung/Desktop/python/" + str(num) + "/OUTPUT1/"  + file_txt     #实例化读取文件地址
    path_write = "C:/Users/samsung/Desktop/python/" + str(num) + "/OUTPUT1/" + str(num) + "PESQ.xlsx"   #实例化生成文件的地址
    with open(path_read,'r') as file:       #打开文件
        lines = file.readlines()            #生成行数
        for line in lines:                  #遍历每一行
            line = line.strip()             #移除每一行首尾指定字符,默认为空格
            line = line.split()             #移除空格进行切片生成新的list
            number = line[1]                #提取信息a
            string = line[2]                #提取信息b
            string_split = string.split("_")      #以下划线_进行拆分
            string_set = set(string_split)        #set去除重复元素,同时set查找更加迅速

            pattern = re.compile(r'\d*(dB)')     #新的匹配关键字---数字dB
            SNR = " "
            for one in string_split:             #遍历每个切片
                match = pattern.match(one)       #寻找dB元素
                # child = os.path.join("%s%s" %(dirpath,onefile))
                if match:
                    SNR = one                    #给SNR赋值
            number = float(number)               #数据转换为浮点数
            if(key_HF in string_set and key_HS not in string_set):        #只存在HF关键字不存在HS关键字
                if(SNR == key_SNR_0dB):#list_HF_0dB                       #对各类分贝情况数据分别进行数据提取
                    list_HF_0dB.append(number)
                if (SNR == key_SNR_5dB):  # list_HF_5dB
                    list_HF_5dB.append(number)
                if (SNR == key_SNR_10dB):  # list_HF_10dB
                    list_HF_10dB.append(number)
                list_HF_alldB.append(number) # list_HF_alldB


            elif(key_HF not in string_set and key_HS in string_set):      #HS情况
                if(SNR == key_SNR_0dB):#list_HS_0dB
                    list_HS_0dB.append(number)
                if (SNR == key_SNR_5dB):  # list_HS_5dB
                    list_HS_5dB.append(number)
                if (SNR == key_SNR_10dB):  # list_HS_10dB
                    list_HS_10dB.append(number)
                list_HS_alldB.append(number) # list_HS_alldB

        list_0dB = list_HF_0dB + list_HS_0dB
        list_5dB = list_HF_5dB + list_HS_5dB
        list_10dB = list_HF_10dB + list_HS_10dB
        list_alldB = list_HF_alldB + list_HS_alldB

        wb = Workbook()        #生成excel工作空间
        ws = wb.active

        ws.append([" ","0dB","5dB","10dB","avg"])
        ws.append(["HF",("%.3f" % np.array(list_HF_0dB).mean()), ("%.3f" % np.array(list_HF_5dB).mean()),   #%.3f保留3位小数
              ("%.3f" % np.array(list_HF_10dB).mean()), ("%.3f" % np.array(list_HF_alldB).mean())])
        ws.append(["HS",("%.3f" % np.array(list_HS_0dB).mean()), ("%.3f" % np.array(list_HS_5dB).mean()),
              ("%.3f" % np.array(list_HS_10dB).mean()), ("%.3f" % np.array(list_HS_alldB).mean())])
        ws.append([" ",("%.3f" % np.array(list_0dB).mean()), ("%.3f" % np.array(list_5dB).mean()),
              ("%.3f" % np.array(list_10dB).mean()), ("%.3f" % np.array(list_alldB).mean())])

        # for row in range(2, 5):
        #     for col in range(1, 6):
        #         ws.cell(row=row, column=col).value =
        wb.save(path_write)

        print(("%.3f" % np.array(list_HF_0dB).mean()),("%.3f" % np.array(list_HF_5dB).mean()),("%.3f" % np.array(list_HF_10dB).mean()),("%.3f" % np.array(list_HF_alldB).mean()))
        print(("%.3f" % np.array(list_HS_0dB).mean()), ("%.3f" % np.array(list_HS_5dB).mean()), ("%.3f" % np.array(list_HS_10dB).mean()),("%.3f" % np.array(list_HS_alldB).mean()))
        #print(("%.3f" % np.array(list_0dB).mean()), ("%.3f" % np.array(list_5dB).mean()), ("%.3f" % np.array(list_10dB).mean()),("%.3f" % np.array(list_alldB).mean()))
    file.close()